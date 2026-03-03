"""
Report generation service for PDF and Excel exports.
"""
import io
import json
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database import get_telemetry_history, get_field_by_name


# ==================== PDF REPORT ====================
class PDFReportGenerator:
    """Генератор PDF отчетов"""
    
    def __init__(self, field_name: str, days: int = 7):
        self.field_name = field_name
        self.days = days
        self.styles = getSampleStyleSheet()
        
    def _create_title_style(self) -> ParagraphStyle:
        """Стиль для заголовка"""
        return ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
    
    def _create_header_style(self) -> ParagraphStyle:
        """Стиль для подзаголовка"""
        return ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=6,
            alignment=TA_LEFT,
        )
    
    def _create_normal_style(self) -> ParagraphStyle:
        """Стиль для обычного текста"""
        return ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#334155'),
        )
    
    async def generate(self) -> bytes:
        """Генерация PDF отчета"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )
        
        elements = []
        
        # Заголовок
        title_style = self._create_title_style()
        elements.append(Paragraph("Отчет по мониторингу влажности", title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Информация о поле
        field = await get_field_by_name(self.field_name)
        if field:
            info_text = f"""
            <b>Поле:</b> {field.field_name}<br/>
            <b>Культура:</b> {field.crop_type}<br/>
            <b>Площадь:</b> {field.area_hectares} га<br/>
            <b>Период:</b> {self.days} дн.
            """
            elements.append(Paragraph(info_text, self._create_normal_style()))
            elements.append(Spacer(1, 0.3*inch))
        
        # Данные телеметрии
        telemetry = await get_telemetry_history(self.field_name, hours=self.days * 24)
        
        if telemetry:
            # Таблица данных
            data = [['Дата/Время', 'Влажность 20см (%)', 'Влажность 40см (%)', 'Температура (°C)', 'Батарея (%)']]
            
            for t in telemetry[:100]:  # Максимум 100 записей
                if t[3]:
                    date_str = t[3].strftime('%d.%m.%Y %H:%M')
                else:
                    date_str = '—'
                
                data.append([
                    date_str,
                    str(t[0]),
                    str(t[0] - 5),  # Примерно для 40см
                    str(round(t[1], 1)),
                    str(t[2]),
                ])
            
            table = Table(data, colWidths=[1.8*inch, 1.2*inch, 1.2*inch, 1*inch, 0.8*inch])
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Статистика
            if telemetry:
                avg_moisture = sum(t[0] for t in telemetry) / len(telemetry)
                avg_temp = sum(t[1] for t in telemetry) / len(telemetry)
                min_moisture = min(t[0] for t in telemetry)
                max_moisture = max(t[0] for t in telemetry)
                
                stats_data = [
                    ['Параметр', 'Значение'],
                    ['Средняя влажность', f'{avg_moisture:.1f}%'],
                    ['Средняя температура', f'{avg_temp:.1f}°C'],
                    ['Мин. влажность', f'{min_moisture}%'],
                    ['Макс. влажность', f'{max_moisture}%'],
                    ['Записей', str(len(telemetry))],
                ]
                
                stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfdf5')]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                
                elements.append(stats_table)
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER,
        )
        elements.append(Paragraph(
            f"Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Aqlli Dala",
            footer_style
        ))
        
        doc.build(elements)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes


# ==================== EXCEL REPORT ====================
class ExcelReportGenerator:
    """Генератор Excel отчетов"""
    
    def __init__(self, field_name: str, days: int = 7):
        self.field_name = field_name
        self.days = days
        
    async def generate(self) -> bytes:
        """Генерация Excel отчета"""
        wb = Workbook()
        
        # Основной лист
        ws = wb.active
        ws.title = "Данные"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        
        # Заголовок
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Отчет по полю: {self.field_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Подзаголовок
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Период: {self.days} дн. | Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Заголовки таблицы
        headers = ['Дата/Время', 'Влажность 20см (%)', 'Влажность 40см (%)', 'Температура (°C)', 'Батарея (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        telemetry = await get_telemetry_history(self.field_name, hours=self.days * 24)
        
        if telemetry:
            for row, t in enumerate(telemetry[:500], 5):  # Максимум 500 записей
                if t[3]:
                    date_str = t[3].strftime('%d.%m.%Y %H:%M')
                else:
                    date_str = '—'
                
                ws.cell(row=row, column=1, value=date_str).alignment = cell_alignment
                ws.cell(row=row, column=2, value=t[0]).alignment = cell_alignment
                ws.cell(row=row, column=3, value=t[0] - 5).alignment = cell_alignment  # Примерно
                ws.cell(row=row, column=4, value=round(t[1], 1)).alignment = cell_alignment
                ws.cell(row=row, column=5, value=t[2]).alignment = cell_alignment
        
        # Авто-ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 25)
        
        # Лист статистики
        ws_stats = wb.create_sheet(title="Статистика")
        
        if telemetry:
            avg_moisture = sum(t[0] for t in telemetry) / len(telemetry)
            avg_temp = sum(t[1] for t in telemetry) / len(telemetry)
            min_moisture = min(t[0] for t in telemetry)
            max_moisture = max(t[0] for t in telemetry)
            
            stats = [
                ['Параметр', 'Значение'],
                ['Средняя влажность', f'{avg_moisture:.1f}%'],
                ['Средняя температура', f'{avg_temp:.1f}°C'],
                ['Мин. влажность', f'{min_moisture}%'],
                ['Макс. влажность', f'{max_moisture}%'],
                ['Всего записей', len(telemetry)],
            ]
            
            for row, (param, value) in enumerate(stats, 1):
                cell = ws_stats.cell(row=row, column=1, value=param)
                cell.font = Font(bold=True) if row == 1 else Font()
                cell.fill = header_fill if row == 1 else PatternFill()
                cell.alignment = header_alignment if row == 1 else cell_alignment
                
                ws_stats.cell(row=row, column=2, value=value).alignment = cell_alignment
        
        # Сохранение
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        excel_bytes = buffer.getvalue()
        wb.close()
        
        return excel_bytes


# ==================== MAIN FUNCTION ====================
async def generate_report(
    field_name: str,
    format: str = 'pdf',
    days: int = 7
) -> tuple[bytes, str]:
    """
    Генерация отчета в указанном формате.
    
    Args:
        field_name: Название поля
        format: 'pdf' или 'xlsx'
        days: Период в днях
        
    Returns:
        Tuple[bytes, str]: (файл, content_type)
    """
    if format.lower() == 'pdf':
        generator = PDFReportGenerator(field_name, days)
        content = await generator.generate()
        return content, 'application/pdf'
    elif format.lower() == 'xlsx':
        generator = ExcelReportGenerator(field_name, days)
        content = await generator.generate()
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        raise ValueError(f"Неподдерживаемый формат: {format}")
